"""
etl_dce.py — Sincronización en Tiempo Real con Data Center Expert
──────────────────────────────────────────────────────────────────────────────
Este módulo se conecta a la API REST de Data Center Expert (DCE) y extrae las
lecturas actuales de los sensores de cada equipo del nodo: Transformador (TR),
Tablero Principal (ML) y los dos Rectificadores (RECT1 y RECT2).

Los datos obtenidos se guardan en el archivo datos_DCE.xlsx, que actúa como
intermediario entre la API y la base de datos MySQL. Desde allí, el módulo
etl.py los carga a las tablas tr_dce, ml_dce y rect_dce.

Flujo completo:
    API DCE → sincronizar_dispositivo_a_excel() → datos_DCE.xlsx → etl.py → MySQL
──────────────────────────────────────────────────────────────────────────────
"""

import pandas as pd
import sys
import os
import time
from datetime import datetime
from openpyxl import load_workbook

# Módulos propios del proyecto
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from src.api_client import GestorDCE
from src.config_dce import DCE_IDS, MAPA_TR, MAPA_ML, MAPA_RECT

# ─────────────────────────────────────────────────────────────
#  RUTA DEL ARCHIVO EXCEL INTERMEDIO
#  Este archivo acumula las lecturas históricas de la API.
#  Cada ejecución agrega una fila nueva sin borrar las anteriores.
# ─────────────────────────────────────────────────────────────
DIR_DATOS         = os.path.join(os.path.dirname(__file__), '..', 'Datos')
ARCHIVO_EXCEL_DCE = os.path.join(DIR_DATOS, 'datos_DCE.xlsx')


def limpiar_valor_dce(valor_str):
    """
    Convierte el valor raw de un sensor de la API a un número o texto limpio.

    La API de DCE retorna los valores como strings con unidades incluidas,
    por ejemplo: "215.5 V", "48.2 A", "Normal", "Online".

    Esta función extrae solo la parte numérica cuando es posible,
    y retorna el texto original cuando el valor es un estado (no numérico).

    Ejemplos:
        "215.5 V"  →  215.5  (float)
        "48.2 A"   →  48.2   (float)
        "Normal"   →  "Normal" (str, estado del equipo)
        "Online"   →  "Online" (str, estado de conexión)
    """
    # Si ya es número, devolverlo sin modificar
    if isinstance(valor_str, (int, float)):
        return valor_str

    texto = str(valor_str).strip()

    try:
        # Intento 1: tomar la parte antes del espacio (ej: "215.5" de "215.5 V")
        parte_numerica = texto.split(' ')[0]
        return float(parte_numerica)
    except ValueError:
        # Intento 2: el valor es texto de estado (ej: "Online", "Float")
        # Se retorna tal cual para no perder la información
        return texto


def guardar_en_excel(df_nuevo, nombre_hoja):
    """
    Agrega una nueva fila al archivo Excel datos_DCE.xlsx sin borrar el historial.

    Este archivo acumula lecturas a lo largo del tiempo. Cada llamada a la API
    agrega una fila nueva con la fecha y hora de la consulta, permitiendo
    conservar el histórico de lecturas operativas.

    Si el archivo no existe, lo crea. Si la hoja no existe, la crea dentro
    del archivo existente. Si la hoja ya existe, concatena los datos nuevos
    con los existentes y sobreescribe la hoja.
    """
    # Caso 1: El archivo no existe → crearlo desde cero
    if not os.path.exists(ARCHIVO_EXCEL_DCE):
        print(f"       Creando archivo nuevo: {ARCHIVO_EXCEL_DCE}")
        with pd.ExcelWriter(ARCHIVO_EXCEL_DCE, engine='openpyxl') as writer:
            df_nuevo.to_excel(writer, sheet_name=nombre_hoja, index=False)
        return

    # Caso 2: El archivo ya existe → agregar datos sin borrar historial
    try:
        book = load_workbook(ARCHIVO_EXCEL_DCE)

        if nombre_hoja in book.sheetnames:
            # La hoja existe: leer todo, concatenar con los datos nuevos y guardar
            df_existente = pd.read_excel(ARCHIVO_EXCEL_DCE, sheet_name=nombre_hoja)
            df_final     = pd.concat([df_existente, df_nuevo], ignore_index=True)

            with pd.ExcelWriter(ARCHIVO_EXCEL_DCE, engine='openpyxl',
                                mode='a', if_sheet_exists='replace') as writer:
                df_final.to_excel(writer, sheet_name=nombre_hoja, index=False)
        else:
            # La hoja no existe: crearla dentro del archivo existente
            with pd.ExcelWriter(ARCHIVO_EXCEL_DCE, engine='openpyxl', mode='a') as writer:
                df_nuevo.to_excel(writer, sheet_name=nombre_hoja, index=False)

    except Exception as e:
        print(f"       Error escribiendo Excel: {e}")


def sincronizar_dispositivo_a_excel(dce, device_id, mapa_columnas, nombre_hoja, extra_data=None):
    """
    Consulta la API de DCE para un dispositivo y guarda los datos en Excel.

    Este es el proceso central del módulo. Para cada equipo del nodo:
        1. Llama a la API usando el ID del dispositivo en DCE.
        2. Recorre los sensores recibidos y los mapea a las columnas del Excel
           usando el diccionario mapa_columnas.
        3. Limpia cada valor con limpiar_valor_dce() para convertir a numérico.
        4. Agrega la fecha y hora de la consulta como columna 'fecha'.
        5. Guarda la fila construida en la hoja correspondiente del Excel.

    Parámetro extra_data: permite agregar columnas adicionales a la fila,
    por ejemplo rectificador_id=1 o rectificador_id=2 para diferenciar
    las lecturas de los dos rectificadores en la misma hoja 'RECT'.
    """
    print(f"    Consultando ID: {device_id} para hoja '{nombre_hoja}'...")

    # PASO 1: Consultar la API
    datos_raw = dce.consultar_equipo(device_id)
    if not datos_raw:
        print("       Sin respuesta del dispositivo.")
        return

    # PASO 2: Construir la fila de datos
    # Se agrega la fecha de consulta y cualquier dato extra (ej: rectificador_id)
    fila = {'fecha': datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
    if extra_data:
        fila.update(extra_data)

    # PASO 3: Mapear sensores de la API a columnas del Excel
    # Se busca cada sensor por su etiqueta (label) en la API y se asigna
    # al nombre de columna correspondiente según el mapa de configuración
    encontrados = 0
    for sensor in datos_raw:
        label_api = sensor.get('label', '')
        valor_api = sensor.get('value', 0)

        for key_map, col_excel in mapa_columnas.items():
            if key_map in label_api:
                fila[col_excel] = limpiar_valor_dce(valor_api)
                encontrados += 1
                break

    # PASO 4: Guardar en Excel si se encontraron datos válidos
    if encontrados > 0:
        df = pd.DataFrame([fila])
        guardar_en_excel(df, nombre_hoja)
        print(f"       Fila agregada a hoja '{nombre_hoja}' ({encontrados} datos).")
    else:
        print("       No se encontraron sensores coincidentes.")


def ejecutar_actualizacion_excel_dce(usuario, password):
    """
    Función principal: conecta con DCE y sincroniza los 4 equipos del nodo.

    Requiere credenciales de acceso a la plataforma Data Center Expert y
    conexión a la VPN corporativa para alcanzar la IP del servidor DCE.

    Orden de sincronización:
        1. TR  (Transformador / Transfer Switch)
        2. ML  (Tablero Principal)
        3. RECT1 (Rectificador 1 — ELTEK SmartPack 2)
        4. RECT2 (Rectificador 2 — ELTEK SmartPack 2)

    Se agrega una pausa de 1 segundo entre consultas para no saturar
    la API del servidor DCE con múltiples peticiones simultáneas.
    """
    IP_DCE = "10.159.125.33"  # IP del servidor Data Center Expert en la red del nodo
    dce    = GestorDCE(IP_DCE, usuario, password)

    print("\n" + "="*60)
    print(" ACTUALIZANDO EXCEL 'datos_DCE.xlsx' DESDE API \n")
    print(f" IP DCE: {IP_DCE} \n")
    print("="*60)

    # Sincronizar Transformador → hoja 'TR'
    sincronizar_dispositivo_a_excel(dce, DCE_IDS["TR"], MAPA_TR, "TR")

    # Sincronizar Tablero ML → hoja 'ML'
    time.sleep(1)
    sincronizar_dispositivo_a_excel(dce, DCE_IDS["ML"], MAPA_ML, "ML")

    # Sincronizar Rectificador 1 → hoja 'RECT' con columna rectificador_id=1
    # Ambos rectificadores se guardan en la misma hoja diferenciados por este ID
    time.sleep(1)
    sincronizar_dispositivo_a_excel(dce, DCE_IDS["RECT1"], MAPA_RECT, "RECT",
                                    extra_data={"rectificador_id": 1})

    # Sincronizar Rectificador 2 → hoja 'RECT' con columna rectificador_id=2
    time.sleep(1)
    sincronizar_dispositivo_a_excel(dce, DCE_IDS["RECT2"], MAPA_RECT, "RECT",
                                    extra_data={"rectificador_id": 2})

    print("\n Excel actualizado correctamente.")
    print(f" Archivo ubicado en: {ARCHIVO_EXCEL_DCE} \n")
